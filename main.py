import requests
import openpyxl
import time
import datetime


GUILD_INVITE_CODE = '' # <- Invite code here 
DISCORD_API_ENDPOINT = f'https://api.mdcdev.me/v1/invites/{GUILD_INVITE_CODE}'
EXCEL_FILE_PATH = 'member_count.xlsx'
WEBHOOK_URL = " " # <- discord webhooh URL





def get_server_boost():
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    server_boost = response_json["guild"]['premiumSubscriptionCount']
    return server_boost

def get_server_feature():
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    server_feature = response_json["guild"]['features']
    return server_feature


def get_member_connected():
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    server_connected = response_json['approximatePresenceCount']
    return server_connected

def get_server_vanity():
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    server_vanity = response_json['guild']['vanityURLCode']
    return server_vanity

def get_member_count():
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    member_count = response_json['approximateMemberCount']
    return member_count



while True: 
    member_count = get_member_count()
    boost_count = get_server_boost()
    member_connected = get_member_connected()
    server_vanity = get_server_vanity()
    server_features = get_server_feature()
    response = requests.get(DISCORD_API_ENDPOINT)
    response_json = response.json()
    print(response_json["guild"]['vanityURLCode'])
    print(f'Member count: {member_count}')

    payload = {
    "username": "Good-uni stalke :D",
    "embeds": [
        {
            "title": "Misc info",
            "description": f"Member count: {member_count}\nMember connected: {member_connected}\nBoost count: {boost_count}\nServer vanity: discord.gg/{server_vanity}",
            "fields": [
                {
                    "name": "Guild feature",
                    "value": f"{server_features}"
                }
            ]
        }
    ]
}

        
    webhook_response = requests.post(WEBHOOK_URL, json=payload)


    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    row_number = sheet.max_row + 1

    sheet.cell(row=row_number, column=1).value = datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S")
    sheet.cell(row=row_number, column=2).value = member_count
    
    workbook.save(EXCEL_FILE_PATH)

    time.sleep(120)
